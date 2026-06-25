import openpyxl, csv, json, re, sys

# ── 1. Members Excel ──────────────────────────────────────────
print("="*60)
print("MEMBERS EXCEL")
print("="*60)
try:
    wb = openpyxl.load_workbook(r'C:\Users\e333638\Desktop\Projetos AntiGravity\COLIH_Salvador\Lista_Membros_16-06-2026.xlsx')
    ws = wb.active
    headers = [ws.cell(1,c).value for c in range(1, ws.max_column+1)]
    print('Headers:', headers)
    print('Total rows:', ws.max_row - 1)
    print('Sample rows:')
    for r in range(2, 5):
        row = {headers[c-1]: ws.cell(r,c).value for c in range(1, ws.max_column+1)}
        print(' ', row)
except Exception as e:
    print('ERROR:', e)

# ── 2. Members CSV ────────────────────────────────────────────
print()
print("="*60)
print("MEMBERS CSV (coordinates)")
print("="*60)
try:
    for enc in ['utf-8-sig', 'latin-1', 'utf-8']:
        try:
            with open(r'C:\Users\e333638\Desktop\Projetos AntiGravity\COLIH_Salvador\lista_membros_para_coordenadas.csv', encoding=enc) as f:
                rows = list(csv.DictReader(f))
            print(f'Encoding: {enc}')
            print('Headers:', list(rows[0].keys()) if rows else 'none')
            print('Total rows:', len(rows))
            print('Sample rows:')
            for row in rows[:5]:
                print(' ', dict(row))
            break
        except UnicodeDecodeError:
            continue
except Exception as e:
    print('ERROR:', e)
