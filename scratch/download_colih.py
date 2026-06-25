import requests
import re
import json

session = requests.Session()
session.headers.update({
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
})

# Step 1: Get CSRF
resp = session.get('https://salvador.colih.med.br', timeout=30)
csrf_match = re.search(r'<input[^>]+name="_token"[^>]+value="([^"]+)"', resp.text)
if not csrf_match:
    csrf_match = re.search(r'name="csrf-token" content="([^"]+)"', resp.text)
csrf_token = csrf_match.group(1)
print(f"CSRF: {csrf_token[:20]}...")

# Step 2: Login
session.post(
    'https://salvador.colih.med.br/auth/login',
    data={'_token': csrf_token, 'username': 'lmc.colihba@gmail.com', 'password': 'Svc@colih26'},
    headers={'Referer': 'https://salvador.colih.med.br/auth/login', 'Content-Type': 'application/x-www-form-urlencoded'},
    allow_redirects=True, timeout=30
)
print("Logged in!")

# Step 3: Download XLS - try different filter combos to get ALL doctors
export_url = 'https://salvador.colih.med.br/panel/doctors/export?type=xls&moreFilters=0&isMultipleRegions=1&cooperationLevel=9&region=1&sortOrder=nm'
print(f"Downloading: {export_url}")

r = session.get(export_url, timeout=60)
ct = r.headers.get('content-type', '')
print(f"Status: {r.status_code}, Content-Type: {ct}, Size: {len(r.content)} bytes")

if r.status_code == 200 and len(r.content) > 1000:
    out = r'C:\Users\e333638\Desktop\Projetos AntiGravity\COLIH_Salvador\colih_export_auto.xls'
    with open(out, 'wb') as f:
        f.write(r.content)
    print(f"Saved to: {out}")
    
    # Try to read it
    try:
        import openpyxl
        wb = openpyxl.load_workbook(out)
        ws = wb.active
        print(f"Rows: {ws.max_row}, Cols: {ws.max_column}")
        headers = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]
        print(f"Headers: {headers}")
    except Exception as e:
        print(f"openpyxl failed: {e}, trying xlrd...")
        # It might be xls (old format)
        print("First bytes:", r.content[:8])
else:
    print("Download failed or empty!")
    print("Response text[:300]:", r.text[:300])

# Also try getting ALL records (no city/region filter) 
print("\n--- Trying ALL regions export ---")
export_all = 'https://salvador.colih.med.br/panel/doctors/export?type=xls&moreFilters=0&isMultipleRegions=1&sortOrder=nm'
r2 = session.get(export_all, timeout=60)
print(f"Status: {r2.status_code}, Content-Type: {r2.headers.get('content-type','')}, Size: {len(r2.content)} bytes")
if r2.status_code == 200 and len(r2.content) > len(r.content):
    out2 = r'C:\Users\e333638\Desktop\Projetos AntiGravity\COLIH_Salvador\colih_export_all.xls'
    with open(out2, 'wb') as f:
        f.write(r2.content)
    print(f"All export saved: {out2}")
